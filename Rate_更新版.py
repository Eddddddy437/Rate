import pandas as pd
import numpy as np
import requests
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- 幣別清單 & 目標月份 ---
currencies = ['USD', 'JPY', 'EUR', 'GBP', 'AUD', 'CAD', 'CNY', 'HKD',
              'SGD', 'CHF', 'SEK', 'ZAR', 'NZD', 'THB', 'PHP', 'IDR', 
              'KRW', 'VND', 'MYR']
month = '2026-03'
all_data = {}

# === 建立 Session（模擬瀏覽器）===
session = requests.Session()
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
    "Connection": "keep-alive"
}

# === 抓取資料（防 503 版本）===
for cur in currencies:
    print(f'抓取 {cur} ...')
    url = f'https://rate.bot.com.tw/xrt/quote/{month}/{cur}'

    try:
        response = session.get(url, headers=headers, timeout=15)

        # 若被擋，自動重試一次
        if response.status_code != 200:
            print(f'⚠️ {cur} retry...')
            time.sleep(3)
            response = session.get(url, headers=headers, timeout=15)

        response.raise_for_status()

        # 交給 pandas 解析 HTML
        bot_data = pd.read_html(response.text)
        df = bot_data[0]

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(1)

        df = df.iloc[:, [0, 2, 1, 3, 4, 5]]
        df.columns = ['掛牌日期', '幣別', '現金買入', '現金賣出', '即期買入', '即期賣出']

        # 修正欄位
        df.loc[:, ['幣別', '現金買入']] = df[['現金買入', '幣別']].values
        df = df.replace('-----------------------', pd.NA)
        df = df.apply(pd.to_numeric, errors='ignore')
        df = df.set_index('掛牌日期')
        df = df.drop(columns=['幣別'])
        df.columns = [f'{cur}_{col}' for col in df.columns]

        all_data[cur] = df

        # 避免被鎖 IP
        time.sleep(1.5)

    except Exception as e:
        print(f'⚠️ {cur} 無法抓取: {e}')

# === 合併所有幣別資料 ===
final_df = pd.concat(all_data.values(), axis=1)

# 每欄位整體平均
avg_row = final_df.mean(numeric_only=True).to_frame().T
avg_row.index = ['3月平均數']

# === 各幣別「買入/賣出平均」 ===
rows = {}
for cur in currencies:
    try:
        cash_cols = [f'{cur}_現金買入', f'{cur}_現金賣出']
        spot_cols = [f'{cur}_即期買入', f'{cur}_即期賣出']
        rows[f'{cur}_現金買入'] = final_df[cash_cols].mean(axis=1, numeric_only=True).mean()
        rows[f'{cur}_現金賣出'] = rows[f'{cur}_現金買入']
        rows[f'{cur}_即期買入'] = final_df[spot_cols].mean(axis=1, numeric_only=True).mean()
        rows[f'{cur}_即期賣出'] = rows[f'{cur}_即期買入']
    except KeyError:
        continue

avg_by_type_row = pd.DataFrame(rows, index=['各幣別買入/賣出平均'])

# 合併到表格底下
final_df = pd.concat([final_df, avg_row, avg_by_type_row])

# === 存成 Excel ===
file_name = 'exrate_FEB.xlsx'
final_df.to_excel(file_name, engine='openpyxl')

# === 用 openpyxl 處理儲存格合併 ===
wb = load_workbook(file_name)
ws = wb.active

# 找到「各幣別買入/賣出平均」所在的 row
target_row = ws.max_row
for cur in currencies:
    try:
        cash_buy_col = list(final_df.columns).index(f'{cur}_現金買入') + 2
        cash_sell_col = list(final_df.columns).index(f'{cur}_現金賣出') + 2
        spot_buy_col = list(final_df.columns).index(f'{cur}_即期買入') + 2
        spot_sell_col = list(final_df.columns).index(f'{cur}_即期賣出') + 2

        ws.merge_cells(start_row=target_row, end_row=target_row,
                       start_column=cash_buy_col, end_column=cash_sell_col)
        ws.merge_cells(start_row=target_row, end_row=target_row,
                       start_column=spot_buy_col, end_column=spot_sell_col)
    except ValueError:
        continue

# === 新增各幣別最後一天買入/賣出平均 ===
ws.append(['各幣別最後一天買入/賣出平均'] + [None]*(len(final_df.columns)-1))
target_row = ws.max_row

for cur in currencies:
    try:
        cash_buy_col = list(final_df.columns).index(f'{cur}_現金買入') + 2
        cash_sell_col = list(final_df.columns).index(f'{cur}_現金賣出') + 2
        spot_buy_col = list(final_df.columns).index(f'{cur}_即期買入') + 2
        spot_sell_col = list(final_df.columns).index(f'{cur}_即期賣出') + 2

        # 現金買入/賣出公式
        ws.cell(row=target_row, column=cash_buy_col, 
                value=f'=AVERAGE({get_column_letter(cash_buy_col)}2:{get_column_letter(cash_sell_col)}2)')
        # 即期買入/賣出公式
        ws.cell(row=target_row, column=spot_buy_col, 
                value=f'=AVERAGE({get_column_letter(spot_buy_col)}2:{get_column_letter(spot_sell_col)}2)')

        # 合併儲存格
        ws.merge_cells(start_row=target_row, end_row=target_row,
                       start_column=cash_buy_col, end_column=cash_sell_col)
        ws.merge_cells(start_row=target_row, end_row=target_row,
                       start_column=spot_buy_col, end_column=spot_sell_col)
    except ValueError:
        continue

wb.save(file_name)
print("✅ 3 月所有幣別匯率已存成 exrate_MAR.xlsx（含平均數 & 各幣別買入/賣出平均 & 各幣別最後一天買入/賣出平均公式，且已合併儲存格）")
